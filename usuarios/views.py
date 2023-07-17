from django.shortcuts import render, redirect, get_object_or_404
from .forms import SignUpForm, LoginForm, AdministradorForm, AlumnoForm, EmpresaForm, RolForm, ProyectoForm, PasswordChangeForm
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from .models import Administrador, Alumno, Empresa, Proyecto, ProyectoAlumno, ConfiguracionPaginas
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db import IntegrityError
from django.http import HttpResponse
import calendar
from openpyxl import Workbook
from datetime import datetime
from xhtml2pdf import pisa
from django.template.loader import get_template
from io import BytesIO

def index(request):
    if request.user.is_authenticated:
        # Si el usuario está autenticado, redirigir según su rol
        if Alumno.objects.filter(user=request.user).exists():
            if ConfiguracionPaginas.objects.first().estado_paginas_alumnos:
                # Redirigir a la página de inicio del alumno
                return redirect('alumno_home')
            else:
                return render(request, 'pagina_desactivada.html')
        elif Empresa.objects.filter(user=request.user).exists():
            if ConfiguracionPaginas.objects.first().estado_paginas_empresas:
                # Redirigir a la página de inicio de la empresa
                return redirect('empresa_home')
            else:
                return render(request, 'pagina_desactivada.html')
        elif Administrador.objects.filter(user=request.user).exists():
            # Redirigir a la página de inicio del administrador
            return redirect('administrador_home')
    else:
        form = LoginForm(request.POST or None)
        msg = None
        if request.method == 'POST':
            if form.is_valid():
                username = form.cleaned_data.get('username')
                password = form.cleaned_data.get('password')
                user = authenticate(username=username, password=password)
                if user is not None and user.es_administrador:
                    login(request, user)
                    return redirect('administrador_home')
                elif user is not None and user.es_alumno:
                    if ConfiguracionPaginas.objects.first().estado_paginas_alumnos:
                        login(request, user)
                        return redirect('alumno_home')
                    else:
                        return render(request, 'pagina_desactivada.html')
                elif user is not None and user.es_empresa:
                    if ConfiguracionPaginas.objects.first().estado_paginas_empresas:
                        login(request, user)
                        return redirect('empresa_home')
                    else:
                        return render(request, 'pagina_desactivada.html')
                else:
                    msg = 'Credenciales inválidas.'
            else:
                msg = 'Ingresa una contraseña.'
        return render(request, 'index.html', {'form': form, 'msg': msg})


def registro(request):
    if request.method == 'POST':
        form = RolForm(request.POST)
        if form.is_valid():
            rol = form.cleaned_data.get('rol')
            if rol == 'alumno':
                return redirect('registro_alumno')
            elif rol == 'empresa':
                return redirect('registro_empresa')
    else:
        form = RolForm()
    return render(request, 'registro.html', {'form': form})


def registro_administrador(request):
    if request.method == 'POST':
        form = AdministradorForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('login_view')
    else:
        form = AdministradorForm()
        user_form = SignUpForm()
    return render(request, 'administrador/registro.html', {'form': form, 'user_form': user_form})


def registro_alumno(request):
    if request.method == 'POST':
        user_form = SignUpForm(request.POST)
        form = AlumnoForm(request.POST)
        if user_form.is_valid() and form.is_valid():
            matricula = form.cleaned_data['matricula']
            user = user_form.save(commit=False)
            user.username = matricula
            user.es_alumno = True
            try:
                user.save()
            except IntegrityError:
                
                user_form.add_error('username', 'La matrícula ya está en uso.')

            if not user_form.errors:
                alumno = form.save(commit=False)
                alumno.user = user
                alumno.save()
                return redirect('index')
    else:
        form = AlumnoForm()
        user_form = SignUpForm()
    return render(request, 'alumno/registro.html', {'form': form, 'user_form': user_form})


def registro_empresa(request):
    if request.method == 'POST':
        user_form = SignUpForm(request.POST)
        form = EmpresaForm(request.POST)
        if user_form.is_valid() and form.is_valid():

            razon_social = form.cleaned_data['razon_social']

            user = user_form.save(commit=False)
            user.username = razon_social
            user.es_empresa = True  # Asigna el rol de empresa al usuario
            user.save()
            empresa = form.save(commit=False)
            empresa.user = user
            empresa.save()
            return redirect('index')
    else:
        form = EmpresaForm()
        user_form = SignUpForm()
    return render(request, 'empresa/registro.html', {'form': form, 'user_form': user_form})


def es_administrador(user):
    return user.es_administrador


def es_alumno(user):
    return user.es_alumno


def es_empresa(user):
    return user.es_empresa


def controlar_paginas(request):
    configuracion = ConfiguracionPaginas.objects.first()

    if request.method == 'POST':
        if 'alumnos' in request.POST:
            configuracion.estado_paginas_alumnos = not configuracion.estado_paginas_alumnos
            messages.success(request, "Páginas de alumnos actualizadas correctamente.")
            configuracion.save()
            return redirect('administrador_alumnos')
        elif 'empresas' in request.POST:
            configuracion.estado_paginas_empresas = not configuracion.estado_paginas_empresas
            messages.success(request, "Páginas de empresas actualizadas correctamente.")
            configuracion.save()
            return redirect('administrador_empresas')
        configuracion.save()

    return redirect('administrador_home')


def pagina_desactivada(request):
    return render(request, 'pagina_desactivada.html')

# Asegura que el usuario esté autenticado


@login_required(login_url='index')
# Verifica si el usuario es un administrador
@user_passes_test(es_administrador)
def administrador_home(request):
    return render(request, 'administrador/home.html')


@login_required(login_url='index')
def logout_view(request):
    logout(request)
    return redirect('index')

# ALUMNO

# Asegura que el usuario esté autenticado


@login_required(login_url='index')
@user_passes_test(es_alumno)  # Verifica si el usuario es un alumno
def alumno_home(request):
    alumno = Alumno.objects.get(user=request.user)
    return render(request, 'alumno/home.html', {'alumno': alumno})


@login_required(login_url='index')
@user_passes_test(es_alumno)  # Verifica si el usuario es un alumno
def alumno_perfil(request):
    alumno = Alumno.objects.get(user=request.user)

    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Contraseña actualizada exitosamente.')
            return redirect('alumno_perfil')
        else:
            messages.error(
                request, 'Por favor corrige los errores en el formulario.')
    else:
        form = PasswordChangeForm(request.user)

    return render(request, 'alumno/perfil.html', {'alumno': alumno, 'form': form})


@login_required(login_url='index')
@user_passes_test(es_alumno)  # Verifica si el usuario es un alumno
def alumno_catalogo(request):
    alumno = Alumno.objects.get(user=request.user)
    programa_educativo = alumno.programa_educativo
    proyectos = Proyecto.objects.filter(
        id_programa_educativo=programa_educativo)
    return render(request, 'alumno/catalogo.html', {'proyectos': proyectos})


@login_required(login_url='index')
@user_passes_test(es_alumno)  # Verifica si el usuario es un alumno
def alumno_proyecto(request):
    user = request.user
    alumno = Alumno.objects.get(user=user)
    proyectos_seleccionados = ProyectoAlumno.objects.filter(
        alumno=alumno).values_list('proyecto', flat=True)
    proyectos = Proyecto.objects.filter(id__in=proyectos_seleccionados)

    context = {
        'proyectos': proyectos
    }

    if request.method == 'POST':
        proyecto_id = request.POST.get('proyecto_id')
        proyecto = Proyecto.objects.get(pk=proyecto_id)
        ProyectoAlumno.objects.filter(
            proyecto=proyecto, alumno=alumno).delete()
        proyecto.vacantes_disponibles += 1
        proyecto.save()
        messages.success(request, "Proyecto deseleccionado exitosamente.")

    return render(request, 'alumno/proyecto.html', context)


# EMPRESA
@login_required(login_url='index')
@user_passes_test(es_empresa)  # Verifica si el usuario es una empresa
def empresa_home(request):
    return render(request, 'empresa/home.html')


@login_required(login_url='index')
@user_passes_test(es_empresa)  # Verifica si el usuario es una empresa
def empresa_perfil(request):
    empresa = Empresa.objects.get(user=request.user)

    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Contraseña actualizada exitosamente.')
            return redirect('empresa_perfil')
        else:
            messages.error(
                request, 'Por favor corrige los errores en el formulario.')
    else:
        form = PasswordChangeForm(request.user)

    return render(request, 'empresa/perfil.html', {'empresa': empresa, 'form': form})


@login_required(login_url='index')
@user_passes_test(es_empresa)  # Verifica si el usuario es una empresa
def empresa_crear_proyecto(request):
    show_confirm_message = True
    show_success_message = False

    if request.method == 'POST':
        if 'confirm' in request.POST:  # Se ha hecho clic en el botón de confirmación
            show_confirm_message = False
        else:
            form = ProyectoForm(request.POST)
            if form.is_valid():

                vacantes = form.cleaned_data['vacantes']

                proyecto = form.save(commit=False)
                proyecto.vacantes_disponibles = vacantes
                proyecto.id_empresa = request.user.empresa
                proyecto.save()
                form = ProyectoForm()  # Reiniciamos el formulario después de guardar el proyecto
                show_confirm_message = False
                show_success_message = True
    else:
        form = ProyectoForm()

    return render(request, 'empresa/crear_proyecto.html', {'form': form, 'show_confirm_message': show_confirm_message, 'show_success_message': show_success_message})


@login_required(login_url='index')
@user_passes_test(es_empresa)  # Verifica si el usuario es una empresa
def empresa_proyectos(request):
    empresa = request.user.empresa
    proyectos = Proyecto.objects.filter(id_empresa=empresa)

    return render(request, 'empresa/proyectos.html', {'proyectos': proyectos})


@login_required(login_url='index')
@user_passes_test(es_alumno)  # Verifica si el usuario es un alumno
def alumno_seleccionar_proyecto(request, proyecto_id):
    # Obtener el proyecto seleccionado
    proyecto = Proyecto.objects.get(pk=proyecto_id)

    # Obtener el usuario alumno actual
    alumno = Alumno.objects.get(user=request.user)

    # Obtener los proyectos seleccionados por el alumno
    proyectos_seleccionados = ProyectoAlumno.objects.filter(alumno=alumno)

    # Obtener los procesos asociados a los proyectos seleccionados
    procesos_seleccionados = proyectos_seleccionados.values_list(
        'proyecto__id_proceso', flat=True)

    # Verificar si el alumno ha alcanzado el límite de proyectos permitidos
    if proyecto.id_proceso.nombre == 'Estancia':
        if procesos_seleccionados.filter(proyecto__id_proceso__nombre='Estancia').count() >= 2:
            # Mostrar un mensaje de error
            messages.error(
                request, "Ya has alcanzado el límite de proyectos de Estancia permitidos.")
            return redirect('alumno_catalogo')
    elif proyecto.id_proceso.nombre == 'Estadía':
        if procesos_seleccionados.filter(proyecto__id_proceso__nombre='Estadía').exists():
            # Mostrar un mensaje de error
            messages.error(
                request, "Ya has seleccionado un proyecto de Estadía.")
            return redirect('alumno_catalogo')
    else:
        # Mostrar un mensaje de error genérico si el proceso no es "Estancia" ni "Estadía"
        messages.error(request, "No puedes seleccionar este proyecto.")
        return redirect('alumno_catalogo')

    # Verificar si hay vacantes disponibles para el proyecto seleccionado
    if proyecto.vacantes_disponibles <= 0:
        # Mostrar un mensaje de error
        messages.error(
            request, "Este proyecto ya no tiene vacantes disponibles.")
        return redirect('alumno_catalogo')

    if ProyectoAlumno.objects.filter(alumno=alumno, proyecto=proyecto).exists():
        # Mostrar un mensaje de error
        messages.error(
            request, "Ya has seleccionado este proyecto previamente.")
        return redirect('alumno_catalogo')

    # Crear una instancia de ProyectoAlumno para asociar el proyecto con el alumno
    ProyectoAlumno.objects.create(proyecto=proyecto, alumno=alumno)

    # Reducir el número de vacantes disponibles en el proyecto
    proyecto.vacantes_disponibles -= 1
    proyecto.save()

    # Mostrar un mensaje de éxito
    messages.success(request, "Proyecto seleccionado exitosamente.")

    # Redireccionar a la página de catálogo de proyectos del alumno
    return redirect('alumno_catalogo')


@login_required(login_url='index')
# Verifica si el usuario es un administrador
@user_passes_test(es_administrador)
def administrador_perfil(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Contraseña actualizada exitosamente.')
            return redirect('administrador_perfil')
        else:
            messages.error(
                request, 'Por favor corrige los errores en el formulario.')
    else:
        form = PasswordChangeForm(request.user)
    return render(request, 'administrador/perfil.html', {'form': form})


# Vista para mostrar la plantilla de administrador/alumnos.html con la lista de alumnos
def administrador_alumnos(request):
    alumnos = Alumno.objects.all()
    configuracion = ConfiguracionPaginas.objects.first()
    if request.method == 'POST':
        user_form = SignUpForm(request.POST)
        form = AlumnoForm(request.POST)
        if user_form.is_valid() and form.is_valid():

            matricula = form.cleaned_data['matricula']

            user = user_form.save(commit=False)
            user.username = matricula
            user.es_alumno = True  # Asigna el rol de alumno al usuario
            user.save()
            alumno = form.save(commit=False)
            alumno.user = user
            alumno.save()

            return redirect('administrador_alumnos')
    else:
        user_form = SignUpForm()
        form = AlumnoForm()

    return render(request, 'administrador/alumnos.html', {'user_form': user_form, 'form': form, 'alumnos': alumnos, 'configuracion': configuracion})


def editar_alumno(request, alumno_id):
    alumno = get_object_or_404(Alumno, pk=alumno_id)
    if request.method == 'POST':
        form = AlumnoForm(request.POST, instance=alumno)
        if form.is_valid():
            form.save()
            return redirect('administrador_alumnos')
    else:
        form = AlumnoForm(instance=alumno)
    return render(request, 'administrador/editar_alumno.html', {'form': form, 'alumno': alumno})


def eliminar_alumno(request, alumno_id):
    alumno = get_object_or_404(Alumno, pk=alumno_id)
    user = alumno.user
    if request.method == 'POST':
        user.delete()
        alumno.delete()
        return redirect('administrador_alumnos')
    return render(request, 'administrador/eliminar_alumno.html', {'alumno': alumno})


@login_required(login_url='index')
# Verifica si el usuario es un administrador
@user_passes_test(es_administrador)
def administrador_empresas(request):
    empresas = Empresa.objects.all()
    configuracion = ConfiguracionPaginas.objects.first()
    if request.method == 'POST':
        user_form = SignUpForm(request.POST)
        form = EmpresaForm(request.POST)
        if user_form.is_valid() and form.is_valid():

            razon_social = form.cleaned_data['razon_social']

            user = user_form.save(commit=False)
            user.username = razon_social
            user.es_empresa = True  # Asigna el rol de empresa al usuario
            user.save()
            empresa = form.save(commit=False)
            empresa.user = user
            empresa.save()
            return redirect('administrador_empresas')
    else:
        form = EmpresaForm()
        user_form = SignUpForm()
    return render(request, 'administrador/empresas.html', {'empresas': empresas, 'form': form, 'user_form': user_form, 'configuracion': configuracion})


def editar_empresa(request, empresa_id):
    empresa = get_object_or_404(Empresa, pk=empresa_id)
    if request.method == 'POST':
        form = EmpresaForm(request.POST, instance=empresa)
        if form.is_valid():
            form.save()
            return redirect('administrador_empresas')
    else:
        form = EmpresaForm(instance=empresa)
    return render(request, 'administrador/editar_empresa.html', {'form': form, 'empresa': empresa})


def eliminar_empresa(request, empresa_id):
    empresa = get_object_or_404(Empresa, pk=empresa_id)
    user = empresa.user
    if request.method == 'POST':
        user.delete()  # Elimina el usuario asociado a la empresa
        empresa.delete()  # Elimina la empresa
        return redirect('administrador_empresas')
    return render(request, 'administrador/eliminar_empresa.html', {'empresa': empresa})

# PROYECTOS


@login_required(login_url='index')
@user_passes_test(es_administrador)
def administrador_proyectos(request):
    proyectos = Proyecto.objects.all()
    if request.method == 'POST':
        form = ProyectoForm(request.POST)
        if form.is_valid():
            proyecto = form.save(commit=False)
            proyecto.vacantes_disponibles = form.cleaned_data['vacantes']
            proyecto.id_empresa = form.cleaned_data['id_empresa']
            proyecto.save()
            form = ProyectoForm()
            return redirect('administrador_proyectos')
    else:
        form = ProyectoForm()
    return render(request, 'administrador/proyectos.html', {'proyectos': proyectos, 'form': form})


def editar_proyecto(request, proyecto_id):
    proyecto = get_object_or_404(Proyecto, pk=proyecto_id)
    if request.method == 'POST':
        form = ProyectoForm(request.POST, instance=proyecto)
        if form.is_valid():
            form.save()
            return redirect('administrador_proyectos')
    else:
        form = ProyectoForm(instance=proyecto)
    return render(request, 'administrador/editar_proyecto.html', {'form': form, 'proyecto': proyecto})


def eliminar_proyecto(request, proyecto_id):
    proyecto = get_object_or_404(Proyecto, pk=proyecto_id)
    if request.method == 'POST':
        proyecto.delete()  # Elimina el proyecto
        return redirect('administrador_proyectos')
    return render(request, 'administrador/eliminar_proyecto.html', {'proyecto': proyecto})


# PROYECTOS SELECCIONADOS
@login_required(login_url='index')
@user_passes_test(es_administrador)
def administrador_proyectos_seleccionados(request):
    proyectos_seleccionados = ProyectoAlumno.objects.all()
    return render(request, 'administrador/proyectos_seleccionados.html', {'proyectos_seleccionados': proyectos_seleccionados})


@login_required(login_url='index')
@user_passes_test(es_administrador)
def eliminar_proyecto_seleccionado(request, proyectoalumno_id):
    proyectoalumno = get_object_or_404(ProyectoAlumno, pk=proyectoalumno_id)
    proyecto = proyectoalumno.proyecto
    if request.method == 'POST':
        proyectoalumno.delete()  # Elimina la relación entre el alumno y el proyecto
        proyecto.vacantes_disponibles += 1  # Incrementa en 1 las vacantes disponibles
        proyecto.save()  # Guarda los cambios en el proyecto
        return redirect('administrador_proyectos_seleccionados')
    return render(request, 'administrador/eliminar_proyecto_seleccionado.html', {'proyectoalumno': proyectoalumno})


# EXPORTAR A EXCEL
def exportar_alumnos(request):

    # Obtén los datos del modelo Empresas
    alumnos = Alumno.objects.all()

    # Crea un nuevo libro de trabajo de Excel
    workbook = Workbook()

    # Obtén la hoja de cálculo activa
    sheet = workbook.active

    # Agrega encabezados de columna
    sheet.append(['Matrícula', 'Nombre Completo', 'Sexo', 'Correo Personal',
                 'Correo Institucional', 'Teléfono', 'Programa Educativo', 'Fecha de registro'])

    # Agrega los datos de las empresas a la hoja de cálculo
    for alumno in alumnos:
        sheet.append([alumno.matricula, f"{alumno.nombre} {alumno.apellido_paterno} {alumno.apellido_materno}", alumno.sexo, alumno.correo_personal, alumno.correo_institucional, alumno.telefono, alumno.programa_educativo.nombre,
                      alumno.user.date_joined.strftime('%Y-%m-%d')])

    # Configura el encabezado de respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # Obtén el nombre del mes en español y el año actual
    mes_actual = datetime.now().month
    nombre_mes = calendar.month_name[mes_actual]
    año_actual = datetime.now().year

    # Genera el nombre de archivo con el mes en español y el año
    nombre_archivo = f'alumnos_{nombre_mes.lower()}{año_actual}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={nombre_archivo}'

    # Guarda el libro de trabajo en la respuesta
    workbook.save(response)

    return response


def exportar_empresas(request):

    # Obtén los datos del modelo Empresas
    empresas = Empresa.objects.all()

    # Crea un nuevo libro de trabajo de Excel
    workbook = Workbook()

    # Obtén la hoja de cálculo activa
    sheet = workbook.active

    # Agrega encabezados de columna
    sheet.append(['Razón Social', 'RFC', 'Teléfono', 'Titular', 'Cargo', 'Correo', 'Dirección',
                 'Nombre Enlace', 'Teléfono Enlace', 'Correo Enlace', 'Fecha de registro'])

    # Agrega los datos de las empresas a la hoja de cálculo
    for empresa in empresas:
        sheet.append([empresa.razon_social, empresa.rfc, empresa.telefono_empresa, empresa.titular, empresa.cargo, empresa.correo,
                      f"{empresa.calle} {empresa.numero}, {empresa.colonia}, {empresa.codigo_postal}, {empresa.ciudad}, {empresa.entidad}", empresa.nombre_enlace, empresa.telefono_enlace, empresa.correo_enlace,
                      empresa.user.date_joined.strftime('%Y-%m-%d')])

    # Configura el encabezado de respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # Obtén el nombre del mes en español y el año actual
    mes_actual = datetime.now().month
    nombre_mes = calendar.month_name[mes_actual]
    año_actual = datetime.now().year

    # Genera el nombre de archivo con el mes en español y el año
    nombre_archivo = f'empresas_{nombre_mes.lower()}{año_actual}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={nombre_archivo}'

    # Guarda el libro de trabajo en la respuesta
    workbook.save(response)

    return response


def exportar_proyectos(request):

    # Obtén los datos del modelo Empresas
    proyectos = Proyecto.objects.all()

    # Crea un nuevo libro de trabajo de Excel
    workbook = Workbook()

    # Obtén la hoja de cálculo activa
    sheet = workbook.active

    # Agrega encabezados de columna
    sheet.append(['Empresa', 'Nombre', 'Periodo', 'Año', 'Programa Educativo', 'Proceso',
                 'Modalidad', 'Asesor Laboral', 'Alumnos Requeridos', 'Objetivo', 'Justificación'])

    # Agrega los datos de las empresas a la hoja de cálculo
    for proyecto in proyectos:
        sheet.append([proyecto.id_empresa.razon_social, proyecto.nombre, proyecto.periodo, proyecto.año, proyecto.id_programa_educativo.nombre,
                     proyecto.id_proceso.nombre, proyecto.modalidad, proyecto.asesor, proyecto.vacantes, proyecto.objetivo, proyecto.justificacion])

    # Configura el encabezado de respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # Obtén el nombre del mes en español y el año actual
    mes_actual = datetime.now().month
    nombre_mes = calendar.month_name[mes_actual]
    año_actual = datetime.now().year

    # Genera el nombre de archivo con el mes en español y el año
    nombre_archivo = f'proyectos_{nombre_mes.lower()}{año_actual}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={nombre_archivo}'

    # Guarda el libro de trabajo en la respuesta
    workbook.save(response)

    return response


def exportar_proyectos_seleccionados(request):

    # Obtén los datos del modelo Empresas
    proyectos_seleccionados = ProyectoAlumno.objects.all()

    # Crea un nuevo libro de trabajo de Excel
    workbook = Workbook()

    # Obtén la hoja de cálculo activa
    sheet = workbook.active

    # Agrega encabezados de columna
    sheet.append(['Matrícula', 'Alumno', 'Empresa', 'Proyecto', 'Programa Educativo',
                 'Proceso', 'Periodo y Año', 'Asesor Laboral', 'Dirección'])

    # Agrega los datos de las empresas a la hoja de cálculo
    for proyectoalumno in proyectos_seleccionados:
        row = [
            proyectoalumno.alumno.matricula,
            f"{proyectoalumno.alumno.nombre} {proyectoalumno.alumno.apellido_paterno} {proyectoalumno.alumno.apellido_materno}",
            proyectoalumno.proyecto.nombre,
            proyectoalumno.proyecto.id_empresa.razon_social,
            proyectoalumno.alumno.programa_educativo.nombre,
            proyectoalumno.proyecto.id_proceso.nombre,
            f"{proyectoalumno.proyecto.periodo} {proyectoalumno.proyecto.año}",
            proyectoalumno.proyecto.asesor,
            f"{proyectoalumno.proyecto.id_empresa.calle} {proyectoalumno.proyecto.id_empresa.numero}, {proyectoalumno.proyecto.id_empresa.colonia}, {proyectoalumno.proyecto.id_empresa.codigo_postal}, {proyectoalumno.proyecto.id_empresa.ciudad}, {proyectoalumno.proyecto.id_empresa.entidad}"
        ]
        sheet.append(row)

    # Configura el encabezado de respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # Obtén el nombre del mes en español y el año actual
    mes_actual = datetime.now().month
    nombre_mes = calendar.month_name[mes_actual]
    año_actual = datetime.now().year

    # Genera el nombre de archivo con el mes en español y el año
    nombre_archivo = f'proyectos_seleccionados_{nombre_mes.lower()}{año_actual}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={nombre_archivo}'

    # Guarda el libro de trabajo en la respuesta
    workbook.save(response)

    return response

def descargar_proyecto_pdf(request, proyecto_id):
    # Obtén el proyecto correspondiente al ID proporcionado
    proyecto = Proyecto.objects.get(id=proyecto_id)

    # Renderiza el template HTML del PDF utilizando los datos del proyecto
    template = get_template('alumno/plantilla_proyecto.html')
    context = {'proyecto': proyecto}
    html = template.render(context)

    # Crea un archivo PDF utilizando la biblioteca xhtml2pdf
    result = BytesIO()
    pdf = pisa.CreatePDF(BytesIO(html.encode("UTF-8")), dest=result)

    # Comprueba si se generó correctamente el PDF
    if not pdf.err:
        # Establece las cabeceras del archivo PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="proyecto_{proyecto.id_proceso.nombre}.pdf"'

        # Copia los contenidos del PDF generado en la respuesta HTTP
        response.write(result.getvalue())
        return response

    return HttpResponse('Error al generar el PDF', status=500)